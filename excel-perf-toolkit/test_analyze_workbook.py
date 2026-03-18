"""Unit tests for analyze_workbook.py"""
import json
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

sys.path.insert(0, str(Path(__file__).parent))
import analyze_workbook as aw


# --- Fixtures ---

@pytest.fixture
def empty_xlsx(tmp_path):
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def clean_xlsx(tmp_path):
    p = tmp_path / "clean.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=f"Item {i}")
        ws.cell(row=i, column=2, value=i * 10)
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def phantom_xlsx(tmp_path):
    p = tmp_path / "phantom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    # Push declared range far beyond data
    ws.cell(row=10000, column=1).font = Font(bold=True)
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def volatile_xlsx(tmp_path):
    p = tmp_path / "volatile.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 61):
        ws.cell(row=i, column=1, value="=NOW()")
        ws.cell(row=i, column=2, value=f"=OFFSET(A{i},0,0)")
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def external_ref_xlsx(tmp_path):
    p = tmp_path / "external.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="='[Other.xlsx]Sheet1'!A1")
    ws.cell(row=2, column=1, value="='\\\\server\\share\\book.xlsx'!B2")
    ws.cell(row=3, column=1, value='=INDIRECT("A1")')
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def named_range_xlsx(tmp_path):
    p = tmp_path / "named.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="data")
    dn1 = DefinedName("BrokenRange", attr_text="#REF!")
    wb.defined_names.add(dn1)
    dn2 = DefinedName("ExtRange", attr_text="='[Budget.xlsx]Sheet1'!A1:D50")
    wb.defined_names.add(dn2)
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def residual_fmt_xlsx(tmp_path):
    p = tmp_path / "residual.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 100):
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = PatternFill(start_color="FFFF00", fill_type="solid")
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def style_bomb_xlsx(tmp_path):
    p = tmp_path / "styles.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 601):
        cell = ws.cell(row=i, column=1, value=f"Item {i}")
        cell.font = Font(size=8 + (i % 30), color=f"{i:06X}")
        cell.fill = PatternFill(start_color=f"{(i * 7) % 0xFFFFFF:06X}", fill_type="solid")
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def bloated_xlsx(tmp_path):
    """Full bloated workbook matching generate_test_workbook.py pathologies."""
    p = tmp_path / "bloated.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Sales Data"
    ws1.append(["Date", "Amount", "Running", "Timestamp"])
    for i in range(2, 52):
        ws1.cell(row=i, column=1, value=f"2025-01-{i-1:02d}")
        ws1.cell(row=i, column=2, value=i * 100)
        ws1.cell(row=i, column=3, value=f"=OFFSET(B{i},0,0)")
        ws1.cell(row=i, column=4, value="=NOW()")
    ws1.cell(row=10000, column=1).font = Font(bold=True)

    ws2 = wb.create_sheet("Lookups")
    ws2.cell(row=1, column=1, value="='[OtherWorkbook.xlsx]Sheet1'!A1")
    ws2.cell(row=2, column=1, value="='\\\\fileserver\\share\\budget.xlsx'!B2")
    ws2.cell(row=3, column=1, value='=INDIRECT("A1")')

    ws3 = wb.create_sheet("Old Import")
    for r in range(1, 500):
        for c in range(1, 20):
            cell = ws3.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, bold=(r % 2 == 0))

    ws4 = wb.create_sheet("Style Bomb")
    for i in range(1, 201):
        cell = ws4.cell(row=i, column=1, value=f"Item {i}")
        cell.font = Font(size=8 + (i % 20), color=f"{i:06X}")

    dn = DefinedName("OldRange", attr_text="#REF!")
    wb.defined_names.add(dn)
    dn2 = DefinedName("ExternalLookup", attr_text="='[Budget2023.xlsx]Summary'!A1:D50")
    wb.defined_names.add(dn2)

    wb.save(p)
    wb.close()
    return p


# --- Helpers ---

def open_and_analyze_sheet(path, index=0, **kwargs):
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        return aw.analyze_sheet(wb.worksheets[index], **kwargs)
    finally:
        wb.close()


# --- Regex Tests ---

class TestVolatileFuncsRegex:
    @pytest.mark.parametrize("formula", [
        "=NOW()", "=TODAY()", "=RAND()", "=RANDBETWEEN(1,10)",
        "=INDIRECT(A1)", "=OFFSET(A1,0,0)", '=INFO("directory")',
        '=CELL("address",A1)', "=now()", "=Now()",
    ])
    def test_matches_volatile(self, formula):
        assert aw.VOLATILE_FUNCS.search(formula)

    @pytest.mark.parametrize("formula", [
        "=SUM(A1:A10)", "=VLOOKUP(A1,B:C,2)", "=IF(A1>0,1,0)",
        "=INDEX(A:A,1)", "=MATCH(A1,B:B,0)", "=LEN(A1)",
    ])
    def test_no_match_nonvolatile(self, formula):
        assert not aw.VOLATILE_FUNCS.search(formula)


class TestExternalRefRegex:
    @pytest.mark.parametrize("formula", [
        "='[Other.xlsx]Sheet1'!A1",
        "='C:\\Users\\file.xlsx'!A1",
        "='\\\\server\\share\\file.xlsx'!A1",
        "=http://example.com/data",
        "=https://example.com/data",
    ])
    def test_matches_external(self, formula):
        assert aw.EXTERNAL_REF.search(formula)

    @pytest.mark.parametrize("formula", [
        "=SUM(A1:A10)", "=Sheet2!A1", '=INDIRECT("A1")',
        "='Sheet Name'!A1",
    ])
    def test_no_match_internal(self, formula):
        assert not aw.EXTERNAL_REF.search(formula)


# --- Zip Analysis ---

class TestAnalyzeZipContents:
    def test_valid_xlsx(self, clean_xlsx):
        result, total = aw.analyze_zip_contents(clean_xlsx)
        assert "error" not in result
        assert result["file_size_bytes"] > 0
        assert result["uncompressed_bytes"] > 0
        assert len(result["top_files"]) > 0

    def test_invalid_zip(self, tmp_path):
        p = tmp_path / "bad.xlsx"
        p.write_text("not a zip file")
        result, total = aw.analyze_zip_contents(p)
        assert "error" in result
        assert total > 0

    def test_all_categories_present(self, clean_xlsx):
        result, _ = aw.analyze_zip_contents(clean_xlsx)
        for key in ("worksheets", "styles", "shared_strings", "vba", "drawings_media", "other"):
            assert key in result["categories"]

    def test_file_size_mb_calculated(self, clean_xlsx):
        result, total = aw.analyze_zip_contents(clean_xlsx)
        assert result["file_size_mb"] == round(total / (1024 * 1024), 2)


# --- Sheet Analysis ---

class TestAnalyzeSheet:
    def test_clean_sheet_no_issues(self, clean_xlsx):
        result = open_and_analyze_sheet(clean_xlsx)
        assert result["data_cell_count"] == 20  # 10 rows x 2 cols
        assert result["formula_count"] == 0
        assert not result["volatile_functions"]
        assert not result["external_references"]
        assert not result["issues"]

    def test_sheet_metadata(self, clean_xlsx):
        result = open_and_analyze_sheet(clean_xlsx)
        assert result["name"] == "Data"
        assert result["declared_max_row"] >= 10
        assert result["declared_max_col"] >= 2

    def test_phantom_range_detected(self, phantom_xlsx):
        result = open_and_analyze_sheet(phantom_xlsx)
        critical = [i for i in result["issues"] if i["severity"] == "critical"]
        assert any("Phantom range" in i["message"] for i in critical)
        assert result["actual_max_row"] == 5
        assert result["actual_max_col"] == 1

    def test_volatile_critical_above_threshold(self, volatile_xlsx):
        result = open_and_analyze_sheet(volatile_xlsx)
        assert result["formula_count"] > 0
        assert len(result["volatile_functions"]) > 0
        vol = [i for i in result["issues"] if "volatile" in i["message"].lower()]
        assert len(vol) == 1
        assert vol[0]["severity"] == "critical"  # >50 volatile calls

    def test_volatile_functions_counted(self, volatile_xlsx):
        result = open_and_analyze_sheet(volatile_xlsx)
        vf = result["volatile_functions"]
        assert "NOW" in vf
        assert "OFFSET" in vf

    def test_external_refs_detected(self, external_ref_xlsx):
        result = open_and_analyze_sheet(external_ref_xlsx)
        assert len(result["external_references"]) > 0
        ext = [i for i in result["issues"] if "external" in i["message"].lower()]
        assert len(ext) >= 1

    def test_empty_sheet_flagged(self, residual_fmt_xlsx):
        result = open_and_analyze_sheet(residual_fmt_xlsx)
        assert result["data_cell_count"] == 0
        msgs = " ".join(i["message"] for i in result["issues"])
        assert "Empty sheet" in msgs or "residual" in msgs.lower()

    def test_style_collection_populates_set(self, style_bomb_xlsx):
        styles = set()
        open_and_analyze_sheet(style_bomb_xlsx, styles=styles)
        assert len(styles) > 100

    def test_no_style_collection_without_param(self, clean_xlsx):
        result = open_and_analyze_sheet(clean_xlsx)
        # Should complete without error; styles=None by default
        assert result["data_cell_count"] == 20

    def test_quick_mode_same_for_small(self, clean_xlsx):
        r1 = open_and_analyze_sheet(clean_xlsx, quick=False)
        r2 = open_and_analyze_sheet(clean_xlsx, quick=True)
        assert r1["data_cell_count"] == r2["data_cell_count"]


# --- Full Workbook Analysis ---

class TestAnalyzeWorkbook:
    def test_clean_no_criticals(self, clean_xlsx):
        r = aw.analyze_workbook(clean_xlsx)
        assert r["summary"]["critical"] == 0
        assert r["summary"]["warning"] == 0

    def test_empty_workbook(self, empty_xlsx):
        r = aw.analyze_workbook(empty_xlsx)
        assert r["summary"]["critical"] == 0
        assert len(r["sheets"]) == 1

    def test_phantom_critical(self, phantom_xlsx):
        r = aw.analyze_workbook(phantom_xlsx)
        assert r["summary"]["critical"] >= 1

    def test_volatile_critical(self, volatile_xlsx):
        r = aw.analyze_workbook(volatile_xlsx)
        assert r["summary"]["critical"] >= 1

    def test_named_range_ref_error(self, named_range_xlsx):
        r = aw.analyze_workbook(named_range_xlsx)
        ref = [i for i in r["issues"] if "#REF!" in i["message"]]
        assert len(ref) == 1
        assert ref[0]["severity"] == "warning"

    def test_named_range_external(self, named_range_xlsx):
        r = aw.analyze_workbook(named_range_xlsx)
        ext = [i for i in r["issues"] if "external source" in i["message"].lower()]
        assert len(ext) >= 1

    def test_style_count_populated(self, style_bomb_xlsx):
        r = aw.analyze_workbook(style_bomb_xlsx)
        assert r["style_count"] > 0

    def test_style_warning_above_threshold(self, style_bomb_xlsx):
        r = aw.analyze_workbook(style_bomb_xlsx)
        style_issues = [i for i in r["issues"] if "style" in i["message"].lower()]
        assert len(style_issues) >= 1

    def test_report_structure(self, clean_xlsx):
        r = aw.analyze_workbook(clean_xlsx)
        for k in ("file", "analyzed_at", "zip_analysis", "sheets", "named_ranges",
                   "external_links", "style_count", "summary", "issues"):
            assert k in r
        for sev in ("critical", "warning", "info"):
            assert sev in r["summary"]

    def test_quick_mode_completes(self, clean_xlsx):
        r = aw.analyze_workbook(clean_xlsx, quick=True)
        assert r["summary"] is not None

    def test_bloated_expected_findings(self, bloated_xlsx):
        r = aw.analyze_workbook(bloated_xlsx)
        assert r["summary"]["critical"] >= 2
        assert r["summary"]["warning"] >= 4

    def test_json_serializable(self, clean_xlsx):
        r = aw.analyze_workbook(clean_xlsx)
        if "top_files" in r.get("zip_analysis", {}):
            r["zip_analysis"]["top_files"] = [
                {"file": f, "size_bytes": s} for f, s in r["zip_analysis"]["top_files"]
            ]
        serialized = json.dumps(r, default=str)
        parsed = json.loads(serialized)
        assert "summary" in parsed

    def test_sheet_issues_counted_in_summary(self, phantom_xlsx):
        r = aw.analyze_workbook(phantom_xlsx)
        sheet_criticals = sum(
            len([i for i in s["issues"] if i["severity"] == "critical"])
            for s in r["sheets"]
        )
        assert r["summary"]["critical"] >= sheet_criticals


# --- Output ---

class TestPrintReport:
    def test_output_contains_header(self, clean_xlsx, capsys):
        aw.print_report(aw.analyze_workbook(clean_xlsx))
        out = capsys.readouterr().out
        assert "EXCEL WORKBOOK ANALYSIS" in out
        assert "SHEET DETAIL" in out

    def test_issues_printed_for_criticals(self, phantom_xlsx, capsys):
        aw.print_report(aw.analyze_workbook(phantom_xlsx))
        out = capsys.readouterr().out
        assert "[CRITICAL]" in out
        assert "FINDINGS" in out

    def test_external_links_section(self, clean_xlsx, capsys):
        r = aw.analyze_workbook(clean_xlsx)
        r["external_links"] = ["\\\\server\\share\\file.xlsx"]
        aw.print_report(r)
        out = capsys.readouterr().out
        assert "EXTERNAL WORKBOOK LINKS" in out

    def test_zip_error_handled(self, clean_xlsx, capsys):
        r = aw.analyze_workbook(clean_xlsx)
        r["zip_analysis"] = {"error": "bad file"}
        aw.print_report(r)
        # Should not crash
        out = capsys.readouterr().out
        assert "EXCEL WORKBOOK ANALYSIS" in out


# --- CLI ---

class TestMain:
    def test_no_args_exits_2(self):
        with patch.object(sys, "argv", ["analyze_workbook.py"]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 2

    def test_missing_file_exits_3(self):
        with patch.object(sys, "argv", ["analyze_workbook.py", "/no/such/file.xlsx"]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 3

    def test_bad_extension_exits_4(self, tmp_path):
        p = tmp_path / "file.csv"
        p.write_text("a,b,c")
        with patch.object(sys, "argv", ["analyze_workbook.py", str(p)]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 4

    def test_clean_exits_0(self, clean_xlsx):
        with patch.object(sys, "argv", ["analyze_workbook.py", str(clean_xlsx)]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 0

    def test_critical_exits_1(self, phantom_xlsx):
        with patch.object(sys, "argv", ["analyze_workbook.py", str(phantom_xlsx)]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 1

    def test_json_output_written(self, clean_xlsx, tmp_path):
        out = tmp_path / "report.json"
        with patch.object(sys, "argv", ["analyze_workbook.py", str(clean_xlsx), "--json", str(out)]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 0
        assert out.exists()
        data = json.loads(out.read_text())
        assert "summary" in data
        assert "file" in data

    def test_json_missing_path_exits_2(self, clean_xlsx):
        with patch.object(sys, "argv", ["analyze_workbook.py", str(clean_xlsx), "--json"]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 2

    def test_quick_flag(self, clean_xlsx):
        with patch.object(sys, "argv", ["analyze_workbook.py", str(clean_xlsx), "--quick"]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 0

    def test_xls_extension_exits_4(self, tmp_path):
        p = tmp_path / "legacy.xls"
        p.write_bytes(b"\x00")
        with patch.object(sys, "argv", ["analyze_workbook.py", str(p)]):
            with pytest.raises(SystemExit) as exc:
                aw.main()
            assert exc.value.code == 4
