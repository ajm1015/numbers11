"""
analyze_workbook.py — Diagnose Excel workbook bloat and performance issues.

Inspects .xlsx/.xlsm files for:
    - Used range vs actual data range (phantom rows/columns)
    - External link references (SMB round-trip triggers)
    - Volatile functions (force full recalc on every edit)
    - Excessive unique cell styles (known Excel perf killer)
    - Named ranges pointing to #REF! or external sources
    - Empty sheets with residual formatting
    - Internal file size breakdown (zip-level analysis)
    - VBA module presence

Usage:
    python analyze_workbook.py <path_to_workbook> [--json output.json]

Exit Codes:
    0 = Analysis complete, no critical issues
    1 = Analysis complete, critical issues found
    2 = Invalid arguments
    3 = File not found or unreadable
    4 = Unsupported file format

Dependencies:
    - Python 3.8+
    - openpyxl >= 3.1.0
"""

import sys
import os
import json
import zipfile
import re
from pathlib import Path
from collections import Counter
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(4)


# --- Constants ---

VOLATILE_FUNCS = re.compile(
    r'\b(INDIRECT|OFFSET|NOW|TODAY|RAND|RANDBETWEEN|INFO|CELL)\s*\(',
    re.IGNORECASE
)
EXTERNAL_REF = re.compile(r"(\[.*?\]|'[A-Z]:\\|'\\\\|https?://)", re.IGNORECASE)

SEVERITY = {'critical': 'CRITICAL', 'warning': 'WARNING', 'info': 'INFO'}

# Thresholds
MAX_STYLES_OK = 500
MAX_STYLES_WARN = 2000
PHANTOM_ROW_RATIO = 5  # used_range > 5x data_range = problem
MAX_EXTERNAL_LINKS_OK = 5


# --- Analysis Functions ---

def analyze_zip_contents(filepath):
    """Break down internal .xlsx zip structure by component size."""
    breakdown = {}
    total = os.path.getsize(filepath)
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            for info in zf.infolist():
                breakdown[info.filename] = info.file_size
    except zipfile.BadZipFile:
        return {'error': 'Not a valid zip/xlsx file'}, total

    # Group by category
    categories = {
        'worksheets': 0, 'styles': 0, 'shared_strings': 0,
        'vba': 0, 'drawings_media': 0, 'other': 0
    }
    for name, size in breakdown.items():
        lower = name.lower()
        if 'worksheet' in lower or '/sheet' in lower:
            categories['worksheets'] += size
        elif 'style' in lower:
            categories['styles'] += size
        elif 'sharedstrings' in lower or 'shared_strings' in lower:
            categories['shared_strings'] += size
        elif 'vba' in lower or 'macro' in lower:
            categories['vba'] += size
        elif any(x in lower for x in ('drawing', 'media', 'image', 'chart')):
            categories['drawings_media'] += size
        else:
            categories['other'] += size

    uncompressed = sum(breakdown.values())
    return {
        'file_size_bytes': total,
        'file_size_mb': round(total / (1024 * 1024), 2),
        'uncompressed_bytes': uncompressed,
        'uncompressed_mb': round(uncompressed / (1024 * 1024), 2),
        'categories': {k: round(v / 1024, 1) for k, v in categories.items()},  # KB
        'top_files': sorted(breakdown.items(), key=lambda x: x[1], reverse=True)[:10]
    }, total


def analyze_sheet(ws, quick=False, styles=None):
    """Analyze a single worksheet for bloat indicators."""
    result = {
        'name': ws.title,
        'declared_dimensions': ws.dimensions,
        'declared_max_row': ws.max_row,
        'declared_max_col': ws.max_column,
        'volatile_functions': [],
        'external_references': [],
        'formula_count': 0,
        'data_cell_count': 0,
        'issues': []
    }

    # Quick mode skips cell-by-cell scan for very large sheets
    cell_limit = 500_000 if quick else 5_000_000
    scanned = 0
    volatile_counter = Counter()
    external_refs = set()
    collect_styles = styles is not None and len(styles) <= MAX_STYLES_WARN

    actual_max_row = 0
    actual_max_col = 0

    for row in ws.iter_rows():
        for cell in row:
            scanned += 1
            if scanned > cell_limit:
                result['issues'].append({
                    'severity': 'info',
                    'message': f'Scan truncated at {cell_limit:,} cells. Sheet may be larger.'
                })
                break

            if collect_styles:
                font, fill = cell.font, cell.fill
                styles.add((
                    str(font.color.rgb if font and font.color else ''),
                    font.size if font else 0,
                    font.bold if font else False,
                    str(fill.fgColor.rgb if fill and fill.fgColor else ''),
                    cell.number_format or '',
                ))
                if len(styles) > MAX_STYLES_WARN:
                    collect_styles = False

            if cell.value is not None:
                result['data_cell_count'] += 1
                actual_max_row = max(actual_max_row, cell.row)
                actual_max_col = max(actual_max_col, cell.column)

            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                formula = str(cell.value)
                result['formula_count'] += 1

                vol_matches = VOLATILE_FUNCS.findall(formula)
                for v in vol_matches:
                    volatile_counter[v.upper()] += 1

                ext_matches = EXTERNAL_REF.findall(formula)
                for e in ext_matches:
                    external_refs.add(f"{cell.coordinate}: {formula[:80]}")
        else:
            continue
        break  # break outer loop if inner broke

    result['actual_max_row'] = actual_max_row
    result['actual_max_col'] = actual_max_col
    result['volatile_functions'] = dict(volatile_counter)
    result['external_references'] = list(external_refs)[:20]  # cap output

    # Phantom range detection
    declared = (ws.max_row or 0) * (ws.max_column or 0)
    actual = actual_max_row * actual_max_col
    if declared > 0 and actual > 0:
        ratio = declared / actual
        if ratio > PHANTOM_ROW_RATIO:
            result['issues'].append({
                'severity': 'critical',
                'message': (
                    f'Phantom range: declared {ws.max_row}x{ws.max_column} '
                    f'but data only fills {actual_max_row}x{actual_max_col} '
                    f'({ratio:.0f}x overextended)'
                )
            })
    elif declared > 0 and actual == 0:
        result['issues'].append({
            'severity': 'warning',
            'message': f'Empty sheet with declared range {ws.dimensions} — residual formatting likely'
        })

    if volatile_counter:
        total_vol = sum(volatile_counter.values())
        result['issues'].append({
            'severity': 'critical' if total_vol > 50 else 'warning',
            'message': f'{total_vol} volatile function calls: {dict(volatile_counter)}'
        })

    if external_refs:
        result['issues'].append({
            'severity': 'critical' if len(external_refs) > MAX_EXTERNAL_LINKS_OK else 'warning',
            'message': f'{len(external_refs)} external reference(s) found in formulas'
        })

    return result


def analyze_workbook(filepath, quick=False):
    """Full workbook analysis. Returns structured report dict."""
    report = {
        'file': str(filepath),
        'analyzed_at': datetime.now().isoformat(),
        'zip_analysis': {},
        'sheets': [],
        'named_ranges': [],
        'external_links': [],
        'style_count': 0,
        'summary': {'critical': 0, 'warning': 0, 'info': 0},
        'issues': []
    }

    # Phase 1: Zip-level analysis
    zip_info, file_size = analyze_zip_contents(filepath)
    report['zip_analysis'] = zip_info

    if file_size > 50 * 1024 * 1024:
        report['issues'].append({
            'severity': 'critical',
            'message': f'File is {zip_info["file_size_mb"]} MB — extremely large for a workbook'
        })
        quick = True  # force quick mode for huge files

    # Phase 2: openpyxl analysis
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=False, keep_vba=True)
    except Exception as e:
        report['issues'].append({'severity': 'critical', 'message': f'Failed to open: {e}'})
        return report

    # Named ranges
    for dn in wb.defined_names.values():
        entry = {'name': dn.name, 'value': str(dn.attr_text)}
        if '#REF!' in str(dn.attr_text):
            entry['issue'] = 'Points to #REF!'
            report['issues'].append({
                'severity': 'warning',
                'message': f'Named range "{dn.name}" points to #REF!'
            })
        elif EXTERNAL_REF.search(str(dn.attr_text)):
            entry['issue'] = 'External reference'
            report['issues'].append({
                'severity': 'warning',
                'message': f'Named range "{dn.name}" references external source'
            })
        report['named_ranges'].append(entry)

    # External links (workbook-level)
    if hasattr(wb, '_external_links'):
        for link in wb._external_links:
            target = getattr(link, 'file_link', str(link))
            report['external_links'].append(str(target))
        if wb._external_links:
            report['issues'].append({
                'severity': 'critical',
                'message': f'{len(wb._external_links)} external workbook link(s) — each triggers SMB/network calls on open'
            })

    # VBA check
    if filepath.suffix.lower() == '.xlsm' or zip_info.get('categories', {}).get('vba', 0) > 0:
        report['issues'].append({
            'severity': 'info',
            'message': 'Workbook contains VBA macros — inspect for auto-execute and external calls'
        })

    # Phase 3: Per-sheet analysis (style counting merged into single cell pass)
    styles = set()
    for ws in wb.worksheets:
        sheet_report = analyze_sheet(ws, quick=quick, styles=styles)
        report['sheets'].append(sheet_report)
        for issue in sheet_report['issues']:
            report['summary'][issue['severity']] += 1

    report['style_count'] = len(styles)
    if len(styles) > MAX_STYLES_WARN:
        report['issues'].append({
            'severity': 'critical',
            'message': f'{len(styles)}+ unique cell styles (>{MAX_STYLES_WARN} is a known perf killer)'
        })
    elif len(styles) > MAX_STYLES_OK:
        report['issues'].append({
            'severity': 'warning',
            'message': f'{len(styles)} unique cell styles (>{MAX_STYLES_OK} can degrade performance)'
        })

    # Count workbook-level issues into summary
    for issue in report['issues']:
        report['summary'][issue['severity']] += 1

    wb.close()
    return report


# --- Output Formatting ---

def print_report(report):
    """Print human-readable report to stdout."""
    SEP = '=' * 72
    print(f"\n{SEP}")
    print(f"  EXCEL WORKBOOK ANALYSIS")
    print(f"  {report['file']}")
    print(f"  {report['analyzed_at']}")
    print(SEP)

    z = report['zip_analysis']
    if 'error' not in z:
        print(f"\n  File Size: {z['file_size_mb']} MB (compressed) / {z['uncompressed_mb']} MB (uncompressed)")
        print(f"  Size Breakdown (KB):")
        for cat, size_kb in sorted(z.get('categories', {}).items(), key=lambda x: x[1], reverse=True):
            if size_kb > 0:
                print(f"    {cat:<20} {size_kb:>10.1f} KB")

    print(f"\n  Sheets: {len(report['sheets'])}")
    print(f"  Named Ranges: {len(report['named_ranges'])}")
    print(f"  External Links: {len(report['external_links'])}")
    print(f"  Unique Styles: {report['style_count']}")

    # Issues summary
    s = report['summary']
    print(f"\n  Issues: {s['critical']} critical, {s['warning']} warning, {s['info']} info")

    # All issues
    all_issues = list(report['issues'])
    for sheet in report['sheets']:
        for issue in sheet['issues']:
            all_issues.append({**issue, 'sheet': sheet['name']})

    if all_issues:
        print(f"\n{'─' * 72}")
        print("  FINDINGS")
        print(f"{'─' * 72}")
        for sev in ('critical', 'warning', 'info'):
            issues = [i for i in all_issues if i['severity'] == sev]
            for issue in issues:
                tag = SEVERITY[sev]
                loc = f" [{issue['sheet']}]" if 'sheet' in issue else ''
                print(f"  [{tag}]{loc} {issue['message']}")

    # Per-sheet detail
    print(f"\n{'─' * 72}")
    print("  SHEET DETAIL")
    print(f"{'─' * 72}")
    for sheet in report['sheets']:
        phantom = ""
        if sheet['actual_max_row'] > 0:
            phantom = f" (actual data: {sheet['actual_max_row']}x{sheet['actual_max_col']})"
        print(f"\n  {sheet['name']}")
        print(f"    Declared range  : {sheet['declared_dimensions']}")
        print(f"    Declared max    : {sheet['declared_max_row']}x{sheet['declared_max_col']}{phantom}")
        print(f"    Data cells      : {sheet['data_cell_count']:,}")
        print(f"    Formulas        : {sheet['formula_count']:,}")
        if sheet['volatile_functions']:
            print(f"    Volatile funcs  : {sheet['volatile_functions']}")
        if sheet['external_references']:
            print(f"    External refs   : {len(sheet['external_references'])} found")
            for ref in sheet['external_references'][:5]:
                print(f"      {ref}")

    # External links detail
    if report['external_links']:
        print(f"\n{'─' * 72}")
        print("  EXTERNAL WORKBOOK LINKS")
        print(f"{'─' * 72}")
        for link in report['external_links']:
            print(f"  {link}")

    print(f"\n{SEP}\n")


# --- Main ---

def main():
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <workbook.xlsx> [--json output.json] [--quick]", file=sys.stderr)
        sys.exit(2)

    filepath = Path(sys.argv[1])
    json_output = None
    quick = '--quick' in sys.argv

    if '--json' in sys.argv:
        idx = sys.argv.index('--json')
        if idx + 1 < len(sys.argv):
            json_output = sys.argv[idx + 1]
        else:
            print("ERROR: --json requires an output path", file=sys.stderr)
            sys.exit(2)

    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}", file=sys.stderr)
        sys.exit(3)

    if filepath.suffix.lower() not in ('.xlsx', '.xlsm'):
        print(f"ERROR: Unsupported format '{filepath.suffix}'. Supports .xlsx and .xlsm only.", file=sys.stderr)
        print("  .xls (legacy binary) requires xlrd. .xlsb requires pyxlsb.", file=sys.stderr)
        sys.exit(4)

    report = analyze_workbook(filepath, quick=quick)
    print_report(report)

    if json_output:
        # Strip non-serializable items from zip top_files tuples
        if 'top_files' in report.get('zip_analysis', {}):
            report['zip_analysis']['top_files'] = [
                {'file': f, 'size_bytes': s} for f, s in report['zip_analysis']['top_files']
            ]
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, default=str)
        print(f"JSON report saved: {json_output}")

    sys.exit(1 if report['summary']['critical'] > 0 else 0)


if __name__ == '__main__':
    main()
