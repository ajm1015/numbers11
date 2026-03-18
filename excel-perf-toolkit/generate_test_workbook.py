"""Generate a test workbook with intentional bloat patterns."""
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

wb = openpyxl.Workbook()

# Sheet 1: Normal data with volatile functions and phantom range
ws1 = wb.active
ws1.title = "Sales Data"
ws1.append(["Date", "Amount", "Running", "Timestamp"])
for i in range(2, 52):
    ws1.cell(row=i, column=1, value=f"2025-01-{i-1:02d}")
    ws1.cell(row=i, column=2, value=i * 100)
    ws1.cell(row=i, column=3, value=f"=OFFSET(B{i},0,0)")  # volatile
    ws1.cell(row=i, column=4, value="=NOW()")  # volatile

# Extend used range artificially (phantom rows)
ws1.cell(row=10000, column=1).font = Font(bold=True)
ws1.cell(row=10000, column=1).value = None  # formatting with no data

# Sheet 2: External references
ws2 = wb.create_sheet("Lookups")
ws2.cell(row=1, column=1, value="='[OtherWorkbook.xlsx]Sheet1'!A1")
ws2.cell(row=2, column=1, value="='\\\\fileserver\\share\\budget.xlsx'!B2")
ws2.cell(row=3, column=1, value="=INDIRECT(\"A1\")")

# Sheet 3: Empty sheet with residual formatting
ws3 = wb.create_sheet("Old Import")
for r in range(1, 500):
    for c in range(1, 20):
        cell = ws3.cell(row=r, column=c)
        cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=(r % 2 == 0))

# Sheet 4: Excessive unique styles
ws4 = wb.create_sheet("Style Bomb")
for i in range(1, 201):
    cell = ws4.cell(row=i, column=1, value=f"Item {i}")
    cell.font = Font(size=8 + (i % 20), color=f"{i:06X}")

# Named range pointing to #REF
from openpyxl.workbook.defined_name import DefinedName
dn = DefinedName("OldRange", attr_text="#REF!")
wb.defined_names.add(dn)
dn2 = DefinedName("ExternalLookup", attr_text="='[Budget2023.xlsx]Summary'!A1:D50")
wb.defined_names.add(dn2)

out = Path(__file__).parent / "test_bloated.xlsx"
wb.save(out)
print(f"Test workbook created: {out} ({out.stat().st_size / 1024:.1f} KB)")
